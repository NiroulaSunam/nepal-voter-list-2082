import requests
from bs4 import BeautifulSoup
import csv, time, sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.utils

BASE_URL    = "https://voterlist.election.gov.np/view_ward.php"
OUTPUT_CSV  = "voter_list.csv"
OUTPUT_XLSX = "voter_list.xlsx"

FORM_PAYLOAD = {
    "state":      "3",
    "district":   "27",
    "vdc_mun":    "5288",
    "ward":       "2",
    "reg_centre": "12475",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Referer":    BASE_URL,
    "Origin":     "https://voterlist.election.gov.np",
    "Content-Type": "application/x-www-form-urlencoded",
}

def parse_table(soup):
    rows = []
    table = soup.find("table", {"id": "tbl_data"})
    if not table:
        return rows
    for tr in table.find("tbody").find_all("tr"):
        tds = tr.find_all("td")
        row = [td.get_text(" ", strip=True) for td in tds[:7]]
        if any(row):
            rows.append(row)
    return rows

def parse_headers(soup):
    table = soup.find("table", {"id": "tbl_data"})
    if not table:
        return []
    ths = table.find("thead").find_all("th")
    h = [th.get_text(strip=True) for th in ths if th.get_text(strip=True)]
    return h[:-1] if h and "विवरण" in h[-1] else h

def scrape():
    s = requests.Session()
    s.headers.update(HEADERS)
    s.verify = False

    print("Submitting form...")
    r = s.post(BASE_URL, data=FORM_PAYLOAD, timeout=180)
    r.encoding = "utf-8"
    soup = BeautifulSoup(r.text, "html.parser")

    if not soup.find("table", {"id": "tbl_data"}):
        print("ERROR: table not found. Page:", soup.get_text()[:300])
        sys.exit(1)

    info = soup.find("div", {"id": "tbl_data_info"})
    if info: print(info.get_text(strip=True))

    col_headers = parse_headers(soup)
    print("Columns:", col_headers)

    all_rows = parse_table(soup)
    print(f"Page 1: {len(all_rows)} rows")

    print(f"\nTotal: {len(all_rows)}")
    return col_headers, all_rows

def save_csv(headers, rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if headers: w.writerow(headers)
        w.writerows(rows)
    print(f"CSV saved -> {path}")

def save_xlsx(headers, rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voter List"
    HF = PatternFill("solid", fgColor="1F4E79")
    AF = PatternFill("solid", fgColor="D6E4F0")
    if headers:
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = HF
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center")
            if ri % 2 == 0: c.fill = AF
    ncols = max(len(headers) if headers else 0, max((len(r) for r in rows), default=0))
    for ci in range(1, ncols+1):
        vals = ([headers[ci-1]] if headers and ci<=len(headers) else []) + \
               [str(r[ci-1]) for r in rows if ci<=len(r)]
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = \
            min(max((len(v) for v in vals), default=10)+4, 45)
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"XLSX saved -> {path}")

if __name__ == "__main__":
    print("=" * 55)
    print("Nepal Voter List — Changunarayan Ward 2 Dubakot")
    print("Expected: 3,889 entries")
    print("=" * 55)
    headers, rows = scrape()
    if rows:
        save_csv(headers, rows, OUTPUT_CSV)
        save_xlsx(headers, rows, OUTPUT_XLSX)
        print(f"\nDone! {len(rows)} records saved.")
    else:
        print("No data found.")